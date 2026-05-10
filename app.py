"""
╔══════════════════════════════════════════════════════════════╗
║  BANKING DEEP INTELLIGENCE — FLASK API v3.0                 ║
║  Chain-of-Thought Extraction | CAMELS | Digital Maturity     ║
╚══════════════════════════════════════════════════════════════╝
"""

import os
import traceback
import json
import time
import hashlib
import csv
import io
import re

from flask import Flask, request, jsonify, render_template
from flask_cors import CORS
from crawler import crawl_website
from extractor import extract_data, create_error_response, get_entity_info, extract_data_single_step
from llm import analyze_strategy, call_ai_api, clean_json, compute_camels_score, compute_digital_maturity

try:
    import PyPDF2
    HAS_PDF = True
except ImportError:
    HAS_PDF = False

try:
    from openpyxl import load_workbook
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False

app = Flask(__name__)
CORS(app)

_cache = {}
CACHE_TTL = 600  # 10 phút


def get_cache_key(urls):
    return hashlib.md5(json.dumps(sorted(urls)).encode()).hexdigest()


def is_cache_valid(entry):
    return time.time() - entry.get("timestamp", 0) < CACHE_TTL


@app.route("/")
def home():
    return render_template("dashboard.html")


@app.route("/health")
def health():
    return {
        "status": "ok",
        "version": "3.0-deep-intelligence",
        "features": [
            "chain-of-thought-extraction",
            "camels-scoring",
            "digital-maturity-model",
            "multi-tier-ai-models",
            "groq-free-tier"
        ]
    }


@app.route("/analyze", methods=["POST"])
def analyze():
    try:
        data = request.get_json(force=True)
        urls = data.get("urls", [])

        if not urls:
            return jsonify({"status": "error", "message": "Vui lòng nhập ít nhất 1 URL"}), 400

        urls = [u.strip() for u in urls if u.strip()][:5]
        urls = [("https://" + u if not u.startswith("http") else u) for u in urls]

        cache_key = get_cache_key(urls)
        if cache_key in _cache and is_cache_valid(_cache[cache_key]):
            print("📦 Returning cached result")
            return jsonify(_cache[cache_key]["data"])

        results = []
        crawl_errors = []

        for idx, url in enumerate(urls, 1):
            try:
                print(f"\n{'='*55}")
                print(f"🔍 [{idx}/{len(urls)}] Processing: {url}")

                raw = crawl_website(url)

                if not raw:
                    entity_info = get_entity_info(url)
                    err = "Không thể crawl website sau nhiều lần thử"
                    print(f"❌ {err}: {url}")
                    crawl_errors.append(f"{url}: {err}")
                    results.append(create_error_response(entity_info, url, err))
                    continue

                print(f"📄 Crawled: {len(raw)} chars")
                extracted = extract_data(raw, url)
                results.append(extracted)

                entity_name = extracted['analysis'].get('entity_name',
                               extracted['analysis'].get('bank_name', 'Unknown'))
                camels = extracted['analysis'].get('camels_scores', {})
                digital = extracted['analysis'].get('digital_maturity_scores', {})
                print(f"✅ {entity_name} | products={len(extracted['analysis'].get('products', []))} | "
                      f"CAMELS={camels.get('overall', '?')} | Digital={digital.get('overall_score', '?')}")

                if idx < len(urls):
                    time.sleep(2)

            except Exception as e:
                print(f"❌ Error processing {url}: {str(e)}")
                traceback.print_exc()
                crawl_errors.append(f"{url}: {str(e)}")
                entity_info = get_entity_info(url)
                results.append(create_error_response(entity_info, url, str(e)))

        valid_results = [r for r in results if r.get("extraction_quality") != "error"]

        if valid_results:
            print(f"\n📊 {len(valid_results)}/{len(results)} valid results. Running master strategy analysis...")
            strategy = analyze_strategy(valid_results)
        else:
            strategy = {
                "executive_summary": (
                    f"Tất cả {len(urls)} websites đều không thể crawl. "
                    f"Lỗi: {'; '.join(crawl_errors[:3])}. "
                    "Vui lòng kiểm tra kết nối mạng hoặc thử lại sau."
                ),
                "competitive_ranking": [],
                "camels_leaderboard": [],
                "strategic_recommendations": {"overall_strategy": "Không có dữ liệu để đưa ra khuyến nghị."}
            }

        meta = {
            "total": len(urls),
            "successful": len(valid_results),
            "errors": len([r for r in results if r.get("extraction_quality") == "error"]),
            "crawl_methods": list(set(r.get("source", "unknown") for r in results)),
            "entity_types": list(set(r.get("entity_type", "unknown") for r in results)),
            "entity_tiers": list(set(r.get("entity_tier", "unknown") for r in results)),
            "crawl_errors": crawl_errors,
            "ai_pipeline": "chain-of-thought-2step",
            "models_used": ["llama-3.3-70b-versatile", "qwen/qwen3-32b", "llama-3.1-8b-instant"],
        }

        response_data = {
            "status": "success",
            "results": results,
            "strategy": strategy,
            "meta": meta,
        }

        _cache[cache_key] = {"data": response_data, "timestamp": time.time()}
        if len(_cache) > 25:
            oldest = min(_cache.keys(), key=lambda k: _cache[k]["timestamp"])
            del _cache[oldest]

        return jsonify(response_data)

    except Exception as e:
        traceback.print_exc()
        return jsonify({"status": "error", "message": str(e)}), 500


def parse_uploaded_file(file_stream, filename):
    """Parse XLSX, CSV, or PDF into structured entity data."""
    ext = filename.lower().split('.')[-1]
    data = []

    try:
        if ext in ['xlsx', 'xls']:
            if not HAS_XLSX:
                return None, "openpyxl chưa được cài đặt"

            wb = load_workbook(file_stream)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))

            if len(rows) < 2:
                return None, "File không đủ dữ liệu (cần ít nhất 2 dòng)"

            header_idx = 0
            header_map = {'entity': 0}
            for i, row in enumerate(rows[:5]):
                row_lower = [str(c).strip().lower() if c else "" for c in row]
                for col_i, cell in enumerate(row_lower):
                    if cell in ['ten_ngan_hang', 'ten_cong_ty', 'ngân hàng', 'công ty',
                                'bank', 'company', 'ngan_hang', 'entity', 'tên', 'name']:
                        header_idx = i
                        header_map['entity'] = col_i
                        break

            for i in range(header_idx + 1, len(rows)):
                row = rows[i]
                if not row or not row[header_map.get('entity', 0)]:
                    continue
                entity_name = str(row[header_map.get('entity', 0)]).strip()
                all_products = []
                for col_idx in range(len(row)):
                    if col_idx == header_map.get('entity', 0):
                        continue
                    if row[col_idx]:
                        products = [p.strip() for p in re.split(r'[;\n,]', str(row[col_idx])) if p.strip()]
                        all_products.extend(products)
                if entity_name:
                    data.append({"ten_to_chuc": entity_name, "san_pham": all_products})

        elif ext == 'csv':
            content = file_stream.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(content))
            for row in reader:
                entity = None
                prods = []
                for key in row:
                    key_lower = key.lower().strip()
                    if key_lower in ['ten_ngan_hang', 'ten_cong_ty', 'ngan_hang', 'cong_ty',
                                     'bank_name', 'company_name', 'bank', 'company',
                                     'entity', 'name', 'tên', 'ngân hàng', 'công ty']:
                        entity = row[key].strip()
                    elif row[key]:
                        products = [p.strip() for p in re.split(r'[;,]', str(row[key])) if p.strip()]
                        prods.extend(products)
                if entity:
                    data.append({"ten_to_chuc": entity, "san_pham": prods})

        elif ext == 'pdf':
            if not HAS_PDF:
                return None, "PyPDF2 chưa được cài đặt"
            reader = PyPDF2.PdfReader(file_stream)
            text = "".join([page.extract_text() or "" for page in reader.pages])

            prompt = (
                f"Extract list of organizations (banks/companies) and products/services. "
                f"Return JSON array: [{{'ten_to_chuc': '...', 'san_pham': ['p1', 'p2']}}]. "
                f"Only return JSON. Text: {text[:6000]}"
            )
            raw = call_ai_api(prompt)
            parsed = clean_json(raw)
            if parsed and isinstance(parsed, list):
                return parsed, None
            return None, "Không thể trích xuất từ PDF"
        else:
            return None, f"Format .{ext} không được hỗ trợ (xlsx, csv, pdf)"

        return (data if data else None), (None if data else "Không tìm thấy dữ liệu trong file")

    except Exception as e:
        return None, f"Lỗi đọc file: {str(e)}"


@app.route('/api/analyze-upload', methods=['POST'])
def analyze_upload():
    if 'file' not in request.files:
        return jsonify({"error": "Không có file được upload"}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({"error": "File rỗng"}), 400

    file_stream = io.BytesIO(file.read())
    items, err = parse_uploaded_file(file_stream, file.filename)

    if err:
        return jsonify({"error": err}), 400
    if not items:
        return jsonify({"error": "Không parse được dữ liệu từ file"}), 400

    prepared = []
    for item in items:
        entity_name = item.get("ten_to_chuc", item.get("ten_ngan_hang", "Unknown"))
        products_raw = item.get("san_pham", item.get("loai_san_pham", []))

        prompt = f"""Phân tích chuyên sâu tổ chức: {entity_name}
Danh sách sản phẩm/dịch vụ: {', '.join(products_raw[:50])}

Là chuyên gia ngân hàng quốc tế, phân tích và trả về JSON:
{{
  "entity_name": "{entity_name}",
  "entity_code": "CODE",
  "entity_type": "bank/company/fintech/insurance",
  "products": [
    {{"category": "CATEGORY", "name": "tên", "features": [], "target": "", "highlight": ""}}
  ],
  "pricing": {{"interest_rates": {{}}, "fees": [], "promotions": [], "pricing_philosophy": ""}},
  "digital_capabilities": [{{"name": "...", "description": ""}}],
  "strategic_analysis": {{
    "positioning": "",
    "target_segments": [],
    "key_differentiators": [],
    "value_proposition": "",
    "growth_vectors": [],
    "strategic_gaps": []
  }},
  "competitive_assessment": {{
    "strengths": [],
    "weaknesses": [],
    "market_position": "",
    "competitive_threat_level": "",
    "unique_selling_points": []
  }},
  "data_confidence": "medium"
}}"""

        try:
            ai_response = call_ai_api(prompt, max_tokens=2500, tier="auto")
            analysis = clean_json(ai_response)
            if not analysis:
                raise Exception("Parse failed")

            analysis["bank_name"] = entity_name
            analysis["bank_code"] = analysis.get("entity_code", entity_name[:3].upper())
            analysis["promotions"] = analysis.get("pricing", {}).get("promotions", [])
            analysis["interest_rates"] = analysis.get("pricing", {}).get("interest_rates", {})

            # Normalize digital_capabilities
            dc = analysis.get("digital_capabilities", [])
            if dc and isinstance(dc[0], dict):
                analysis["digital_capabilities_detailed"] = dc
                analysis["digital_capabilities"] = [d.get("name", "") for d in dc if d.get("name")]

            # Compute scores
            analysis["camels_scores"] = compute_camels_score(analysis)
            analysis["digital_maturity_scores"] = compute_digital_maturity(analysis)

        except Exception as e:
            print(f"⚠️ AI analysis failed for {entity_name}: {e}")
            analysis = {
                "entity_name": entity_name,
                "entity_code": entity_name[:4].upper(),
                "entity_type": "company",
                "bank_name": entity_name,
                "bank_code": entity_name[:4].upper(),
                "products": [
                    {"name": p, "category": guess_category(p), "features": [], "target": "", "highlight": ""}
                    for p in products_raw
                ],
                "pricing": {"promotions": [], "fees": [], "interest_rates": {}},
                "promotions": [],
                "interest_rates": {},
                "digital_capabilities": [],
                "digital_capabilities_detailed": [],
                "strategic_analysis": {
                    "positioning": "", "target_segments": [], "key_differentiators": []
                },
                "competitive_assessment": {
                    "strengths": [], "weaknesses": [], "market_position": ""
                },
                "data_confidence": "low",
            }
            analysis["camels_scores"] = compute_camels_score(analysis)
            analysis["digital_maturity_scores"] = compute_digital_maturity(analysis)

        entity_type = analysis.get("entity_type", "company")
        n_products = len(analysis.get("products", []))
        quality = "good" if n_products >= 5 else "limited"

        prepared.append({
            "url": "",
            "analysis": analysis,
            "extraction_quality": quality,
            "source": "file_upload",
            "entity_type": entity_type,
            "entity_tier": "File Upload",
        })

    strategy = analyze_strategy(prepared)
    return jsonify({
        "status": "success",
        "results": prepared,
        "strategy": strategy,
        "meta": {
            "total": len(prepared),
            "successful": len(prepared),
            "errors": 0,
            "source": "file_upload",
            "ai_pipeline": "deep-analysis-v3",
        }
    })


def guess_category(product_name):
    """Đoán category từ tên sản phẩm"""
    n = product_name.lower()
    if any(w in n for w in ['tiết kiệm', 'gửi', 'savings', 'deposit', 'tích lũy']):
        return 'SAVINGS'
    if any(w in n for w in ['vay', 'tín dụng', 'loan', 'credit', 'cho vay', 'bnpl']):
        return 'LOAN'
    if any(w in n for w in ['thẻ', 'card', 'visa', 'mastercard', 'debit', 'jcb', 'amex']):
        return 'CARD'
    if any(w in n for w in ['app', 'mobile', 'digital', 'internet', 'online', 'ebank', 'số', 'ekyc']):
        return 'DIGITAL'
    if any(w in n for w in ['bảo hiểm', 'insurance', 'bảo vệ', 'nhân thọ', 'sức khỏe']):
        return 'INSURANCE'
    if any(w in n for w in ['đầu tư', 'quỹ', 'trái phiếu', 'investment', 'fund', 'gold', 'vàng', 'chứng khoán']):
        return 'INVESTMENT'
    if any(w in n for w in ['thanh toán', 'payment', 'chuyển tiền', 'transfer', 'qr', 'nạp tiền']):
        return 'PAYMENT'
    if any(w in n for w in ['doanh nghiệp', 'business', 'corporate', 'sme', 'trade']):
        return 'BUSINESS'
    if any(w in n for w in ['wealth', 'private banking', 'quản lý tài sản', 'hnw']):
        return 'WEALTH'
    return 'OTHER'


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))
    app.run(host="0.0.0.0", port=port, debug=False)
